from itertools import zip_longest
from datetime import datetime

from openpyxl import load_workbook

from django.conf import settings

from reports.models import FileData, Report
from common.models import Sex


FILE_DATA_NAMES = {
    "Пол пациента": "sex",
    "Дата рождения пациента": "birthday",
    "ID пациента": "patient_external_id",
    "Код МКБ-10": "diagnosis_code",
    "Диагноз": "diagnosis_name",
    "Дата оказания услуги": "date_service",
    "Должность": "doctor_position",
    "Назначения": "medical_appointments",
    "ФИО врача": "doctor_name",
    "Код услуги": "service_code",
    "ID врача": "doctor_external_id",
}


class FileDataParser:

    def __init__(self, file_path: str, report: Report):
        self.file_path = file_path
        self.report = report

    def is_fill_row(self, row: tuple) -> bool:
        return any(row)

    def proccess_field(self, field: str, value: str) -> str | None:
        """Обработка значений в вид, который принимает БД"""
        if value is None:
            return value

        if isinstance(value, str):
            value = value.strip()

        if field == "sex":
            return Sex.MALE if value.strip() == "Муж" else Sex.FEMALE
        if field in ("birthday", "date_service"):
            date_obj = datetime.strptime(value, "%d.%m.%Y")
            return date_obj.strftime("%Y-%m-%d")

        return value

    def parse_row(self, row: tuple, names: tuple) -> dict[str, int | str | None]:
        data = dict()
        for cell in zip_longest(row, names):
            field_name = FILE_DATA_NAMES.get(cell[1])
            if field_name:
                value = self.proccess_field(field_name, cell[0])
                data[field_name] = value

        return data

    def run(self) -> None:
        file_path = settings.MEDIA_ROOT / self.file_path
        wb = load_workbook(file_path, read_only=True)
        ws = wb.active

        # Берем названия колонок в первой строке
        names = next(ws.iter_rows(values_only=True))

        for row in ws.iter_rows(values_only=True, min_row=2):
            if self.is_fill_row(row):
                data = self.parse_row(row, names)
                FileData.objects.create(**data, report=self.report)
