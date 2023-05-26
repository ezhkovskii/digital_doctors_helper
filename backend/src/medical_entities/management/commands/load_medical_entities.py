from django.core.management.base import BaseCommand
from django.db import transaction
from openpyxl.reader.excel import load_workbook

from medical_entities.models import Diagnosis, MedicalAppointments


class Command(BaseCommand):

    def _create_diagnoses(self, diagnoses):
        diagnoses_objects = []
        if diagnoses:
            diagnoses = diagnoses.split(";")
            for d in diagnoses:
                if d:
                    d_strip = d.strip()
                    diagnosis = Diagnosis.objects.filter(code=d_strip).first()
                    if not diagnosis:
                        diagnoses_objects.append(Diagnosis.objects.create(code=d_strip))
                    else:
                        diagnoses_objects.append(diagnosis)

        return diagnoses_objects

    def _create_ma(self, ma_name, ma_syn, ma_req, ma_code):
        syn = []
        if ma_syn:
            syn = ma_syn.split(";")
            syn = [s.strip() for s in syn]
        if ma_req:
            req = True if ma_req == 'да' else False
        else:
            raise Exception('не заполнено required')

        if ma_code is None:
            ma_code = ""

        ma = MedicalAppointments.objects.create(name=ma_name, synonyms=syn, required=req, service_code=ma_code)

        return ma

    def is_fill_row(self, row: tuple) -> bool:
        return any(row)

    def handle(self, *args, **options):
        file_path = "medical.xlsx"
        wb = load_workbook(file_path, read_only=True)
        ws = wb.active

        with transaction.atomic():
            for row in ws.iter_rows(values_only=True, min_row=2):
                if self.is_fill_row(row):
                    ma_name = row[0]
                    ma_syn = row[1]
                    ma_req = row[2]
                    ma_code = row[3]
                    diagnoses = row[4]

                    diagnoses_objects = self._create_diagnoses(diagnoses)

                    ma = self._create_ma(ma_name, ma_syn, ma_req, ma_code)
                    for d in diagnoses_objects:
                        ma.diagnoses.add(d)
